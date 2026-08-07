from datetime import date
from io import BytesIO

from fastapi import APIRouter, Request
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates

from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.database.database import SessionLocal
from app.models.kas_masuk import KasMasuk
from app.models.kas_keluar import KasKeluar


router = APIRouter()

templates = Jinja2Templates(
    directory="app/templates"
)


# =========================================================
# NAMA BULAN
# =========================================================

NAMA_BULAN = [
    "",
    "Januari",
    "Februari",
    "Maret",
    "April",
    "Mei",
    "Juni",
    "Juli",
    "Agustus",
    "September",
    "Oktober",
    "November",
    "Desember"
]


# =========================================================
# HELPER LOGIN
# =========================================================

def cek_login(request: Request):

    if "user_id" not in request.session:
        return None

    return request.session


# =========================================================
# HELPER ROLE
# =========================================================

def cek_role(user):

    return user["role"] in [
        "ketua",
        "wakil",
        "bendahara"
    ]


# =========================================================
# HALAMAN LAPORAN KEUANGAN
# =========================================================

@router.get("/laporan")
def halaman_laporan(
    request: Request,
    bulan: int | None = None,
    tahun: int | None = None
):

    # =====================================================
    # CEK LOGIN
    # =====================================================

    user = cek_login(request)

    if not user:

        return RedirectResponse(
            "/login",
            status_code=302
        )


    # =====================================================
    # CEK ROLE
    # =====================================================

    if not cek_role(user):

        return RedirectResponse(
            "/",
            status_code=302
        )


    db: Session = SessionLocal()

    try:

        # =================================================
        # TANGGAL SEKARANG
        # =================================================

        sekarang = date.today()


        # =================================================
        # FILTER DEFAULT
        # =================================================

        if bulan is None:
            bulan = sekarang.month

        if tahun is None:
            tahun = sekarang.year


        # =================================================
        # VALIDASI BULAN
        # =================================================

        if bulan < 1 or bulan > 12:
            bulan = sekarang.month


        bulan_terpilih = NAMA_BULAN[bulan]


        # =================================================
        # KAS MASUK PERIODE
        # =================================================

        kas_masuk_list = (
            db.query(KasMasuk)
            .options(
                joinedload(KasMasuk.anggota)
            )
            .filter(
                func.extract(
                    "month",
                    KasMasuk.tanggal
                ) == bulan,

                func.extract(
                    "year",
                    KasMasuk.tanggal
                ) == tahun
            )
            .order_by(
                KasMasuk.tanggal.desc(),
                KasMasuk.id.desc()
            )
            .all()
        )


        # =================================================
        # KAS KELUAR PERIODE
        # =================================================

        kas_keluar_list = (
            db.query(KasKeluar)
            .filter(
                func.extract(
                    "month",
                    KasKeluar.tanggal
                ) == bulan,

                func.extract(
                    "year",
                    KasKeluar.tanggal
                ) == tahun
            )
            .order_by(
                KasKeluar.tanggal.desc(),
                KasKeluar.id.desc()
            )
            .all()
        )


        # =================================================
        # TOTAL KAS MASUK PERIODE
        # =================================================

        total_masuk_periode = (
            db.query(
                func.coalesce(
                    func.sum(KasMasuk.nominal),
                    0
                )
            )
            .filter(
                func.extract(
                    "month",
                    KasMasuk.tanggal
                ) == bulan,

                func.extract(
                    "year",
                    KasMasuk.tanggal
                ) == tahun
            )
            .scalar()
        )


        # =================================================
        # TOTAL KAS KELUAR PERIODE
        # =================================================

        total_keluar_periode = (
            db.query(
                func.coalesce(
                    func.sum(KasKeluar.nominal),
                    0
                )
            )
            .filter(
                func.extract(
                    "month",
                    KasKeluar.tanggal
                ) == bulan,

                func.extract(
                    "year",
                    KasKeluar.tanggal
                ) == tahun
            )
            .scalar()
        )


        # =================================================
        # TOTAL SELURUH KAS MASUK
        # =================================================

        total_masuk_semua = (
            db.query(
                func.coalesce(
                    func.sum(KasMasuk.nominal),
                    0
                )
            )
            .scalar()
        )


        # =================================================
        # TOTAL SELURUH KAS KELUAR
        # =================================================

        total_keluar_semua = (
            db.query(
                func.coalesce(
                    func.sum(KasKeluar.nominal),
                    0
                )
            )
            .scalar()
        )


        # =================================================
        # SALDO KESELURUHAN
        # =================================================

        saldo_kas = (
            total_masuk_semua
            - total_keluar_semua
        )


        # =================================================
        # SELISIH PERIODE
        # =================================================

        selisih_periode = (
            total_masuk_periode
            - total_keluar_periode
        )


        # =================================================
        # DAFTAR TAHUN
        # =================================================

        tahun_awal = sekarang.year - 5
        tahun_akhir = sekarang.year + 1

        daftar_tahun = list(
            range(
                tahun_akhir,
                tahun_awal - 1,
                -1
            )
        )


        # =================================================
        # TEMPLATE
        # =================================================

        return templates.TemplateResponse(
            request=request,
            name="laporan.html",
            context={

                "user": user,

                "kas_masuk_list":
                    kas_masuk_list,

                "kas_keluar_list":
                    kas_keluar_list,

                "total_masuk_periode":
                    total_masuk_periode,

                "total_keluar_periode":
                    total_keluar_periode,

                "total_masuk_semua":
                    total_masuk_semua,

                "total_keluar_semua":
                    total_keluar_semua,

                "saldo_kas":
                    saldo_kas,

                "selisih_periode":
                    selisih_periode,

                "bulan":
                    bulan,

                "tahun":
                    tahun,

                "bulan_terpilih":
                    bulan_terpilih,

                "nama_bulan":
                    NAMA_BULAN,

                "daftar_tahun":
                    daftar_tahun,
            }
        )

    finally:

        db.close()


# =========================================================
# DOWNLOAD LAPORAN EXCEL
# =========================================================

@router.get("/laporan/excel")
def download_laporan_excel(
    request: Request,
    bulan: int | None = None,
    tahun: int | None = None
):

    # =====================================================
    # CEK LOGIN
    # =====================================================

    user = cek_login(request)

    if not user:

        return RedirectResponse(
            "/login",
            status_code=302
        )


    # =====================================================
    # CEK ROLE
    # =====================================================

    if not cek_role(user):

        return RedirectResponse(
            "/",
            status_code=302
        )


    # =====================================================
    # FILTER DEFAULT
    # =====================================================

    sekarang = date.today()

    if bulan is None:
        bulan = sekarang.month

    if tahun is None:
        tahun = sekarang.year


    # =====================================================
    # VALIDASI BULAN
    # =====================================================

    if bulan < 1 or bulan > 12:
        bulan = sekarang.month


    bulan_terpilih = NAMA_BULAN[bulan]


    db: Session = SessionLocal()

    try:

        # =================================================
        # DATA KAS MASUK
        # =================================================

        kas_masuk_list = (
            db.query(KasMasuk)
            .options(
                joinedload(KasMasuk.anggota)
            )
            .filter(
                func.extract(
                    "month",
                    KasMasuk.tanggal
                ) == bulan,

                func.extract(
                    "year",
                    KasMasuk.tanggal
                ) == tahun
            )
            .order_by(
                KasMasuk.tanggal.asc(),
                KasMasuk.id.asc()
            )
            .all()
        )


        # =================================================
        # DATA KAS KELUAR
        # =================================================

        kas_keluar_list = (
            db.query(KasKeluar)
            .filter(
                func.extract(
                    "month",
                    KasKeluar.tanggal
                ) == bulan,

                func.extract(
                    "year",
                    KasKeluar.tanggal
                ) == tahun
            )
            .order_by(
                KasKeluar.tanggal.asc(),
                KasKeluar.id.asc()
            )
            .all()
        )


        # =================================================
        # TOTAL PERIODE
        # =================================================

        total_masuk_periode = sum(
            item.nominal
            for item in kas_masuk_list
        )

        total_keluar_periode = sum(
            item.nominal
            for item in kas_keluar_list
        )

        selisih_periode = (
            total_masuk_periode
            - total_keluar_periode
        )


        # =================================================
        # TOTAL KESELURUHAN
        # =================================================

        total_masuk_semua = (
            db.query(
                func.coalesce(
                    func.sum(KasMasuk.nominal),
                    0
                )
            )
            .scalar()
        )

        total_keluar_semua = (
            db.query(
                func.coalesce(
                    func.sum(KasKeluar.nominal),
                    0
                )
            )
            .scalar()
        )

        saldo_kas = (
            total_masuk_semua
            - total_keluar_semua
        )


        # =================================================
        # BUAT WORKBOOK
        # =================================================

        workbook = Workbook()


        # =================================================
        # STYLE
        # =================================================

        font_judul = Font(
            bold=True,
            size=16
        )

        font_subjudul = Font(
            bold=True,
            size=12
        )

        font_header = Font(
            bold=True
        )

        align_tengah = Alignment(
            horizontal="center",
            vertical="center"
        )

        align_kiri = Alignment(
            horizontal="left",
            vertical="center"
        )

        garis_tipis = Side(
            style="thin"
        )

        border = Border(
            left=garis_tipis,
            right=garis_tipis,
            top=garis_tipis,
            bottom=garis_tipis
        )


        # =================================================
        # SHEET 1 - RINGKASAN
        # =================================================

        ws_ringkasan = workbook.active

        ws_ringkasan.title = "Ringkasan"


        ws_ringkasan.merge_cells(
            "A1:D1"
        )

        ws_ringkasan["A1"] = (
            "LAPORAN KEUANGAN KARANG TARUNA SENDANGAN"
        )

        ws_ringkasan["A1"].font = font_judul
        ws_ringkasan["A1"].alignment = align_tengah


        ws_ringkasan.merge_cells(
            "A2:D2"
        )

        ws_ringkasan["A2"] = (
            f"Periode {bulan_terpilih} {tahun}"
        )

        ws_ringkasan["A2"].font = font_subjudul
        ws_ringkasan["A2"].alignment = align_tengah


        # =================================================
        # DATA RINGKASAN
        # =================================================

        data_ringkasan = [

            (
                "Total Kas Masuk Periode",
                total_masuk_periode
            ),

            (
                "Total Kas Keluar Periode",
                total_keluar_periode
            ),

            (
                "Selisih Periode",
                selisih_periode
            ),

            (
                "Total Seluruh Kas Masuk",
                total_masuk_semua
            ),

            (
                "Total Seluruh Kas Keluar",
                total_keluar_semua
            ),

            (
                "Saldo Kas Keseluruhan",
                saldo_kas
            ),
        ]


        baris = 4

        for label, nominal in data_ringkasan:

            ws_ringkasan.cell(
                row=baris,
                column=1,
                value=label
            )

            ws_ringkasan.cell(
                row=baris,
                column=2,
                value=nominal
            )

            ws_ringkasan.cell(
                row=baris,
                column=2
            ).number_format = (
                '"Rp" #,##0'
            )

            ws_ringkasan.cell(
                row=baris,
                column=1
            ).border = border

            ws_ringkasan.cell(
                row=baris,
                column=2
            ).border = border

            baris += 1


        ws_ringkasan.column_dimensions["A"].width = 32
        ws_ringkasan.column_dimensions["B"].width = 20
        ws_ringkasan.column_dimensions["C"].width = 5
        ws_ringkasan.column_dimensions["D"].width = 5


        # =================================================
        # SHEET 2 - KAS MASUK
        # =================================================

        ws_masuk = workbook.create_sheet(
            "Kas Masuk"
        )


        ws_masuk.merge_cells(
            "A1:H1"
        )

        ws_masuk["A1"] = (
            f"KAS MASUK - {bulan_terpilih} {tahun}"
        )

        ws_masuk["A1"].font = font_judul
        ws_masuk["A1"].alignment = align_tengah


        header_masuk = [
            "No",
            "Tanggal",
            "Nama Anggota",
            "Bulan",
            "Tahun",
            "Nominal",
            "Metode",
            "Keterangan"
        ]


        for kolom, header in enumerate(
            header_masuk,
            start=1
        ):

            cell = ws_masuk.cell(
                row=3,
                column=kolom,
                value=header
            )

            cell.font = font_header
            cell.alignment = align_tengah
            cell.border = border


        # =================================================
        # ISI KAS MASUK
        # =================================================

        row = 4

        for nomor, kas in enumerate(
            kas_masuk_list,
            start=1
        ):

            nama_anggota = "-"

            if kas.anggota:
                nama_anggota = kas.anggota.nama


            data = [
                nomor,
                kas.tanggal,
                nama_anggota,
                kas.bulan,
                kas.tahun,
                kas.nominal,
                kas.metode,
                kas.keterangan or "-"
            ]


            for kolom, value in enumerate(
                data,
                start=1
            ):

                cell = ws_masuk.cell(
                    row=row,
                    column=kolom,
                    value=value
                )

                cell.border = border
                cell.alignment = align_kiri


            # Format tanggal
            ws_masuk.cell(
                row=row,
                column=2
            ).number_format = "dd-mm-yyyy"


            # Format nominal
            ws_masuk.cell(
                row=row,
                column=6
            ).number_format = '"Rp" #,##0'


            row += 1


        # =================================================
        # TOTAL KAS MASUK
        # =================================================

        ws_masuk.cell(
            row=row,
            column=5,
            value="TOTAL"
        )

        ws_masuk.cell(
            row=row,
            column=5
        ).font = font_header


        ws_masuk.cell(
            row=row,
            column=6,
            value=total_masuk_periode
        )

        ws_masuk.cell(
            row=row,
            column=6
        ).font = font_header

        ws_masuk.cell(
            row=row,
            column=6
        ).number_format = '"Rp" #,##0'


        # =================================================
        # LEBAR KOLOM KAS MASUK
        # =================================================

        lebar_masuk = [
            7,
            15,
            30,
            15,
            12,
            18,
            15,
            35
        ]

        for index, width in enumerate(
            lebar_masuk,
            start=1
        ):

            ws_masuk.column_dimensions[
                get_column_letter(index)
            ].width = width


        ws_masuk.freeze_panes = "A4"


        # =================================================
        # SHEET 3 - KAS KELUAR
        # =================================================

        ws_keluar = workbook.create_sheet(
            "Kas Keluar"
        )


        ws_keluar.merge_cells(
            "A1:F1"
        )

        ws_keluar["A1"] = (
            f"KAS KELUAR - {bulan_terpilih} {tahun}"
        )

        ws_keluar["A1"].font = font_judul
        ws_keluar["A1"].alignment = align_tengah


        header_keluar = [
            "No",
            "Tanggal",
            "Kategori",
            "Keperluan",
            "Nominal",
            "Keterangan"
        ]


        for kolom, header in enumerate(
            header_keluar,
            start=1
        ):

            cell = ws_keluar.cell(
                row=3,
                column=kolom,
                value=header
            )

            cell.font = font_header
            cell.alignment = align_tengah
            cell.border = border


        # =================================================
        # ISI KAS KELUAR
        # =================================================

        row = 4

        for nomor, kas in enumerate(
            kas_keluar_list,
            start=1
        ):

            data = [
                nomor,
                kas.tanggal,
                kas.kategori,
                kas.keperluan,
                kas.nominal,
                kas.keterangan or "-"
            ]


            for kolom, value in enumerate(
                data,
                start=1
            ):

                cell = ws_keluar.cell(
                    row=row,
                    column=kolom,
                    value=value
                )

                cell.border = border
                cell.alignment = align_kiri


            ws_keluar.cell(
                row=row,
                column=2
            ).number_format = "dd-mm-yyyy"


            ws_keluar.cell(
                row=row,
                column=5
            ).number_format = '"Rp" #,##0'


            row += 1


        # =================================================
        # TOTAL KAS KELUAR
        # =================================================

        ws_keluar.cell(
            row=row,
            column=4,
            value="TOTAL"
        )

        ws_keluar.cell(
            row=row,
            column=4
        ).font = font_header


        ws_keluar.cell(
            row=row,
            column=5,
            value=total_keluar_periode
        )

        ws_keluar.cell(
            row=row,
            column=5
        ).font = font_header

        ws_keluar.cell(
            row=row,
            column=5
        ).number_format = '"Rp" #,##0'


        # =================================================
        # LEBAR KOLOM KAS KELUAR
        # =================================================

        lebar_keluar = [
            7,
            15,
            20,
            35,
            18,
            35
        ]

        for index, width in enumerate(
            lebar_keluar,
            start=1
        ):

            ws_keluar.column_dimensions[
                get_column_letter(index)
            ].width = width


        ws_keluar.freeze_panes = "A4"


        # =================================================
        # SIMPAN KE MEMORY
        # =================================================

        output = BytesIO()

        workbook.save(output)

        output.seek(0)


        # =================================================
        # NAMA FILE
        # =================================================

        nama_file = (
            f"Laporan_Keuangan_"
            f"{bulan_terpilih}_"
            f"{tahun}.xlsx"
        )


        # =================================================
        # DOWNLOAD
        # =================================================

        return StreamingResponse(

            output,

            media_type=(
                "application/"
                "vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),

            headers={

                "Content-Disposition":
                    f'attachment; filename="{nama_file}"'

            }
        )

    finally:

        db.close()